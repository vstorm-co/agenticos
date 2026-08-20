"""File upload service."""

import io
import logging
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.exceptions import BadRequestError, NotFoundError
from app.db.models.chat_file import ChatFile
from app.repositories import chat_file as chat_file_repo
from app.services.file_storage import (
    ALLOWED_MIME_TYPES,
    classify_file,
    get_file_storage,
)

logger = logging.getLogger(__name__)

PREVIEW_LINES = 3
"""How many lines of a file the upload response carries back.

Enough for a composer card to show what was attached — the first rows of a CSV,
the opening of a pasted stack trace — and no more. This is a label, not a
reader: the file itself is one click away in the preview panel.
"""

PREVIEW_CHARS = 240
"""A second bound, for a file whose three lines are one long line each."""


def make_preview(parsed_content: str | None) -> str | None:
    """The head of a file's extracted text, for a client to render beside its name.

    Derived here rather than in the browser because the browser cannot derive it:
    a PDF or a DOCX is bytes until this service has parsed it, and the client only
    ever holds an id and a filename once the upload has answered. Returning it
    with the upload is also the only version that survives a redraw, where a
    client-side excerpt would be a second source of truth about the same file.

    `None` for anything with no text — an image, or a parse that failed — so a
    card renders its thumbnail or its name alone rather than an empty quote.
    """
    if not parsed_content:
        return None
    head = "\n".join(parsed_content.splitlines()[:PREVIEW_LINES])[:PREVIEW_CHARS].strip()
    return head or None


class FileUploadService:
    """Service for file upload validation, parsing, and persistence."""

    ALLOWED_MIME_TYPES = ALLOWED_MIME_TYPES

    def __init__(self, db: AsyncSession):
        self.db = db

    @staticmethod
    def validate_upload(content_type: str | None, size: int) -> tuple[bool, str | None]:
        """Validate a chat attachment's type and size.

        The ceiling is `CHAT_MAX_UPLOAD_SIZE_MB`, and it is a setting because it
        was a literal: `MAX_UPLOAD_SIZE` in `file_storage.py`, 10 MiB, which no
        operator could raise while `/health` published the knowledge base's 50
        and the composer checked against that. A 20MB attachment passed the
        client check, was read into memory, crossed the wire in full and was
        refused here by a number no configuration produced (#498).

        It is a *different* setting from the knowledge base's rather than the
        same one, because the two surfaces fail differently at the same size: a
        document is chunked and read back through retrieval, while an attachment
        to an agent with no workspace is pasted whole into the prompt.

        Returns:
            Tuple of (is_valid, error_message).
        """
        if content_type not in ALLOWED_MIME_TYPES:
            return False, f"File type '{content_type}' is not supported."
        limit_mb = settings.CHAT_MAX_UPLOAD_SIZE_MB
        if size > limit_mb * 1024 * 1024:
            return False, f"File too large. Maximum size is {limit_mb}MB."
        return True, None

    @staticmethod
    def classify_file(mime_type: str, filename: str) -> str:
        """Classify file type based on MIME type and extension."""
        return classify_file(mime_type, filename)

    async def parse_content(
        self,
        data: bytes,
        file_type: str,
        mime_type: str = "",
    ) -> str | None:
        """Parse file content based on file type.

        Returns extracted text content or None if parsing fails.
        """
        if file_type == "text":
            return self._parse_text_content(data, mime_type)
        if file_type == "pdf":
            return self._parse_pdf_content(data)
        if file_type == "docx":
            return self._parse_docx_content(data)
        if file_type == "spreadsheet":
            return self._parse_spreadsheet_content(data)
        return None

    @staticmethod
    def _parse_text_content(data: bytes, mime_type: str) -> str | None:
        """Extract text content from text-based files."""
        try:
            return data.decode("utf-8")
        except (UnicodeDecodeError, ValueError):
            return None

    @staticmethod
    def _parse_pdf_pymupdf(data: bytes) -> str | None:
        """Extract text from PDF using PyMuPDF."""
        try:
            import pymupdf

            doc: Any = pymupdf.open(stream=data, filetype="pdf")  # type: ignore[no-untyped-call]
            text_parts = []
            for page in doc:
                text = page.get_text("text")
                if text.strip():
                    text_parts.append(text.strip())
            doc.close()
            return "\n\n".join(text_parts) if text_parts else None
        except Exception as e:
            logger.warning("PyMuPDF PDF parsing failed: %s", e)
            return None

    def _parse_pdf_content(self, data: bytes) -> str | None:
        """Read a PDF attached to a chat message.

        PyMuPDF, and only PyMuPDF. A chat attachment belongs to no collection,
        so there is no stored configuration to read a parser choice from, and
        the two alternatives were worse than nothing here: LlamaParse bills per
        page and needs a key, LiteParse needs a heavier local toolchain, and
        both were wrapped in `except Exception: return self._parse_pdf_pymupdf()`
        - which meant this file had been silently using PyMuPDF all along. The
        LiteParse branch could not have worked at all: it called a `parse_async`
        method the binding does not define.
        """
        return self._parse_pdf_pymupdf(data)

    @staticmethod
    def _parse_docx_content(data: bytes) -> str | None:
        """Extract text from DOCX."""
        try:
            from docx import Document as DOCXDocument

            doc: Any = DOCXDocument(io.BytesIO(data))
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except Exception as e:
            logger.warning("DOCX parsing failed: %s", e)
            return None

    @staticmethod
    def _parse_spreadsheet_content(data: bytes) -> str | None:
        """Extract a workbook as tab-separated rows, one block per sheet.

        Every sheet, named. A workbook's second sheet is where the data usually
        is - the first is a cover or an index - and a reader that took only the
        active one would answer questions about a file it had half read.

        Tabs rather than commas: a cell holding "1,5" is a number in half of
        Europe, and comma-separating those rows produces a table with a column
        that appears and disappears down the page. `read_only` because a workbook
        is opened here to be read once and thrown away, and it is what keeps a
        large one from being materialised in full.

        Trailing empty cells and empty rows are dropped. A sheet whose used range
        is wider than its data - which is most of them, after a column has been
        cleared - otherwise contributes rows of tabs, and those cost tokens to say
        nothing.
        """
        try:
            from openpyxl import load_workbook

            workbook: Any = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            try:
                blocks: list[str] = []
                for sheet in workbook.worksheets:
                    rows: list[str] = []
                    for row in sheet.iter_rows(values_only=True):
                        cells = ["" if value is None else str(value) for value in row]
                        while cells and cells[-1] == "":
                            cells.pop()
                        if cells:
                            rows.append("\t".join(cells))
                    if rows:
                        blocks.append(f"Sheet: {sheet.title}\n" + "\n".join(rows))
                return "\n\n".join(blocks) or None
            finally:
                # `read_only` keeps file handles open until it is closed, and this
                # runs inside a request.
                workbook.close()
        except Exception as e:
            logger.warning("Spreadsheet parsing failed: %s", e)
            return None

    async def upload(
        self,
        *,
        user_id: Any,
        file_data: bytes,
        filename: str,
        content_type: str | None,
    ) -> ChatFile:
        """Validate, parse, persist, and record a chat file upload.

        Raises:
            BadRequestError: If file type or size is invalid.
        """
        is_valid, error = self.validate_upload(content_type, len(file_data))
        if not is_valid:
            raise BadRequestError(message=error or "Invalid file")

        file_type = self.classify_file(content_type or "", filename)
        parsed_content = await self.parse_content(file_data, file_type, content_type or "")

        storage = get_file_storage()
        storage_path = await storage.save(str(user_id), filename, file_data)

        return await self.create_chat_file(
            user_id=user_id,
            filename=filename,
            mime_type=content_type or "application/octet-stream",
            size=len(file_data),
            storage_path=storage_path,
            file_type=file_type,
            parsed_content=parsed_content,
        )

    async def discard(self, chat_file: ChatFile) -> None:
        """Delete a stored file and the row recording it - the inverse of `upload`.

        For a caller whose turn was refused after the bytes were already stored.
        The bytes go first: a row deleted while the file survives leaves nothing
        pointing at it.
        """
        await get_file_storage().delete(chat_file.storage_path)
        await chat_file_repo.delete(self.db, db_file=chat_file)

    def get_file_path(self, storage_path: str) -> str | None:
        """Resolve a storage path to an absolute filesystem path."""
        full_path = get_file_storage().get_full_path(storage_path)
        return str(full_path) if full_path is not None else None

    async def get_user_file(self, file_id: Any, user_id: Any) -> ChatFile:
        """Get a file by ID, verifying ownership.

        Raises:
            NotFoundError: If file does not exist or user has no access.
        """
        chat_file = await chat_file_repo.get_by_id(self.db, file_id)
        if not chat_file or str(chat_file.user_id) != str(user_id):
            raise NotFoundError(message="File not found")
        return chat_file

    async def create_chat_file(
        self,
        *,
        user_id: Any,
        filename: str,
        mime_type: str,
        size: int,
        storage_path: str,
        file_type: str,
        parsed_content: str | None = None,
    ) -> ChatFile:
        """Create a chat file record in the database."""
        return await chat_file_repo.create(
            self.db,
            user_id=user_id,
            filename=filename,
            mime_type=mime_type,
            size=size,
            storage_path=storage_path,
            file_type=file_type,
            parsed_content=parsed_content,
        )
